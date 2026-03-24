import re
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.exceptions import InvalidFileException

from app.schemas.parameters import IEDFilesData, OAFilesData, CurrentParameter, ReferenceParameter
from app.exceptions import (
    EmptyFileError,
    InvalidFileFormatError,
    InvalidFileContentError,
    IEDNotIdentifiedError,
    FileProcessingError,
    FileTooLargeError
)

FORBIDDEN_WORDS: set[str] = {"RELE", "RELES", "RELAY", "DO", "DA", "DE", "TYPE", "MODELO"}

OA_METADATA_TERMS: set[str] = {
    "ORDEM DE AJUSTE",
    "DADOS DA INSTALAÇÃO",
    "DATA DE EMISSÃO",
    "DATA DE VENCIMENTO",
    "RELE",
    "TRAFO MARCA",
    "EQUIPAMENTO",
    "COMPONENTE",
    "SUBESTAÇÃO",
    "CONTROLE DE TAP E TEMPERATURA",
}

IED_PARAMETER_PATTERN = re.compile(r'([A-Za-z0-9_\-]+)\s*[:=,]\s*"?([^"\r\n#]+)"?')
IED_IGNORED_KEYS: set[str] = {"RELAYTYPE", "BFID", "PARTNO", "RID", "TID", "INFO", "DID"}

MAX_XLSX_SIZE_BYTES: int = 5 * 1024 * 1024

# FORMATAÇÃO DOS NOMES DOS IEDS (MARCA MODELO)
def format_relay_name(name: str | None) -> str:
    if not name:
        return "NÃO IDENTIFICADO"

    clean_name = (
        str(name)
        .strip()
        .upper()
        .replace("-", " ")
        .replace("_", " ")
        .replace('"', "")
        .replace("'", "")
    )

    words = [w for w in clean_name.split() if w not in FORBIDDEN_WORDS]

    if not words:
        return "NÃO IDENTIFICADO"

    return " ".join(words[:2])


def metadata_row(param_value: str | None, row_text: str) -> bool:
    if param_value:
        param_upper = param_value.upper()
        if any(term in param_upper for term in OA_METADATA_TERMS):
            return True

    return "DATA DE EMISSÃO" in row_text or "DADOS DA INSTALAÇÃO" in row_text

# EXTRAÇÃO DOS PARÂMETROS DA ORDEM DE AJUSTE
def parse_oa_parameters(ws: Worksheet, filename: str) -> list[ReferenceParameter]:
    parameters_list = []
    current_group = "GENERAL SETTINGS"

    col_param = col_desc = col_range = col_setting = None

    data_start_row = 0
    all_rows = list(ws.iter_rows(values_only=True))

    for i, row in enumerate(all_rows):
        for idx, cell in enumerate(row):
            if not cell: continue
            value = str(cell).strip().upper()

            if value in ["PARAMETRO", "COMANDO"]:
                col_param = idx
            elif value in ["DESCRICAO", "DESCRIÇÃO", "AJUSTES GLOBAIS"]:
                col_desc = idx
            elif value == "FAIXA DE AJUSTE":
                col_range = idx
            elif value == "AJUSTE SUGERIDO":
                col_setting = idx

        if col_param is not None and col_setting is not None:
            data_start_row = i
            break
    
    if col_param is None or col_setting is None:
        raise InvalidFileContentError(
            filename=filename,
            details="Não localizadas as colunas obrigatórias 'PARAMETRO' e 'AJUSTE SUGERIDO'."
        )
    

    if data_start_row > 0:
        prev_row = all_rows[data_start_row - 1]
        for cell in prev_row:
            if cell and str(cell).strip():
                current_group = str(cell).strip()
                break


    for i in range(data_start_row + 1, len(all_rows)):
        row = all_rows[i]

        row_full_text = " ".join([str(c).upper() for c in row if c])
        
        # ÚLTIMO BLOCO DA PLANILHA
        if "ELABORAÇÃO" in row_full_text or "ELABORACAO" in row_full_text:
            break
        
        raw_param = row[col_param] if len(row) > col_param else None
        param_value = str(raw_param).strip() if raw_param else None

        if metadata_row(param_value, row_full_text):
            continue

        raw_setting = row[col_setting] if len(row) > col_setting else None
        setting_value = str(raw_setting).strip() if raw_setting is not None else None
        
        if param_value and param_value.upper() in ["PARAMETRO", "COMANDO"]:
            prev_row = all_rows[i - 1]
            for cell in prev_row:
                if cell and str(cell).strip():
                    current_group = str(cell).strip()
                    break
            continue 

        if (not param_value and not setting_value) or (param_value and not setting_value):
            found_text = next((str(c).strip() for c in row if c and str(c).strip().upper() != "NONE"), None)
            if found_text and found_text.upper() not in ["AJUSTES GLOBAIS", "DESCRIÇÃO"]:
                current_group = found_text
            continue

        if param_value and setting_value:
            if setting_value.upper() in ["AJUSTE SUGERIDO", "VALOR", "AJUSTE"]:
                continue

            desc = str(row[col_desc]).strip() if col_desc is not None and row[col_desc] else ""
            range_value = str(row[col_range]).strip() if col_range is not None and row[col_range] else ""

            parameters_list.append(
                ReferenceParameter(
                    group=current_group,
                    parameter=param_value,
                    description=desc,
                    setting_range=range_value,
                    reference_value=setting_value,
                )
            )
        
    if not parameters_list:
        raise InvalidFileContentError(
            filename=filename,
            details="Nenhum conteúdo válido foi extraído do arquivo"
        )
        
    return parameters_list

# LEITURA E EXTRAÇÃO DO ARQUIVO DA ORDEM DE AJUSTE
def parse_oa_file(content: bytes, filename: str) -> OAFilesData:
    if not content:
        raise EmptyFileError(filename=filename)
    
    if len(content) > MAX_XLSX_SIZE_BYTES:
        raise FileTooLargeError(
            filename=filename,
            max_size_mb=int(MAX_XLSX_SIZE_BYTES / (1024 * 1024))
        )
    
    try:
        workbook = load_workbook(filename=BytesIO(content), data_only=True)

        if not workbook.worksheets:
            raise InvalidFileContentError(filename=filename, details="O arquivo Excel não possui conteúdo.")
        
        sheet = workbook.worksheets[0]

        relay_model = format_relay_name(sheet["D3"].value)
        substation = str(sheet["D6"].value).strip() if sheet["D6"].value else "NÃO IDENTIFICADA"
        component_name = str(sheet["D7"].value).strip() if sheet["D7"].value else "NÃO IDENTIFICADO"

        params = parse_oa_parameters(sheet, filename)
        workbook.close()

        return OAFilesData(
            filename=filename,
            substation=substation,
            component_name=component_name,
            relay_model=relay_model,
            parameters=params
        )

    except InvalidFileException:
        raise InvalidFileFormatError(filename=filename, accepted_formats=".xlsx")
    except Exception as e:
        if isinstance(e, (EmptyFileError, InvalidFileFormatError, InvalidFileContentError)):
            raise e
        raise FileProcessingError(filename=filename, details=str(e))



# EXTRAÇÃO DOS PARÂMETROS DA IED
def parse_ied_parameters(text: str) -> list[CurrentParameter]:
    param_list = []

    clean_text = "".join(ch for ch in text if ch.isprintable() or ch in "\n\r\t")

    for key, value in IED_PARAMETER_PATTERN.findall(clean_text):
        upper_key = key.upper()
        if upper_key in IED_IGNORED_KEYS:
            continue
        
        clean_value = value.split("#")[0].strip().replace('"', '').replace("'", "")

        if clean_value:
            param_list.append(CurrentParameter(
                parameter=upper_key,
                current_value=clean_value
            ))

    return param_list

# LEITURA E EXTRAÇÃO DO ARQUIVO DA IED
def parse_ied_file(content: bytes, filename: str) -> IEDFilesData:
    if not content:
        raise EmptyFileError(filename=filename)

    try:
        text = content.decode("utf-8")
    except UnicodeDecodeError:
        text = content.decode("latin-1", errors="ignore")
    
    relay_match = re.search(r"(?:FID|RELAYTYPE)\s*[:=,]\s*([^\s,\r\n]+)", text, re.I)
    
    if not relay_match:
        relay_model = "MODELO DESCONHECIDO"
    else:
        relay_model = format_relay_name(relay_match.group(1))

    params = parse_ied_parameters(text)

    if not params:
        raise InvalidFileContentError(
            filename=filename, 
            details="Nenhum parâmetro detectado no formato CHAVE,\"VALOR\"."
        )

    return IEDFilesData(
        filename=filename,
        relay_model=relay_model,
        parameters=params
    )