from app.schemas import ArquivoIED, ArquivoOA, ParametroAtual, ParametroReferencia
import re
from openpyxl import load_workbook
from io import BytesIO

# FORMATAÇÃO DOS NOMES DOS IEDS (MARCA MODELO)
PALAVRAS_PROIBIDAS: set[str] = {"RELE", "RELES", "RELAY", "DO", "DA", "DE", "TYPE", "MODELO"}

TERMOS_PROIBIDOS_OA: set[str] = {
    "ORDEM DE AJUSTE",
    "DADOS DA INSTALAÇÃO",
    "DATA DE EMISSÃO",
    "DATA DE VENCIMENTO",
    "RELE",
    "TRAFO MARCA",
    "EQUIPAMENTO",
    "SUBESTAÇÃO:",
    "CONTROLE DE TAP E TEMPERATURA",
}

PADRAO_PARAMETRO_IED = re.compile(r'([A-Za-z0-9]+),"([^"]*)"')

CHAVES_IGNORADAS_IED: set[str] = {"RELAYTYPE", "BFID", "PARTNO", "RID", "TID", "INFO", "DID"}

MAX_TAMANHO_XLSX_BYTES: int = 10 * 1024 * 1024

def formatar_nome(nome: str) -> str:
    if not nome:
        return "NÃO IDENTIFICADO"

    nome_limpo = (
        str(nome)
        .strip()
        .upper()
        .replace("-", " ")
        .replace("_", " ")
        .replace('"', "")
        .replace("'", "")
    )

    palavras_uteis = [p for p in nome_limpo.split() if p not in PALAVRAS_PROIBIDAS]

    if not palavras_uteis:
        return "NÃO IDENTIFICADO"

    # Retorna no máximo marca + modelo (duas palavras)
    return " ".join(palavras_uteis[:2])

# EXTRAÇÃO DOS PARÂMETROS DA ORDEM DE AJUSTE
def linha_metadados(valor_param: str | None, texto_linha: str) -> bool:
    if valor_param:
        param_upper = valor_param.upper()
        if any(termo in param_upper for termo in TERMOS_PROIBIDOS_OA):
            return True

    if "DATA DE EMISSÃO" in texto_linha or "DADOS DA INSTALAÇÃO" in texto_linha:
        return True

    return False

def leitura_parametros_oa(ws) -> list[ParametroReferencia]:
    lista_parametros = []
    grupo_atual = "GENERAL SETTINGS"

    col_param = None
    col_desc = None
    col_faixa = None
    col_ajuste = None

    linha_inicio_dados = 0
    todas_linhas = list(ws.iter_rows(values_only=True))

    for i, linha in enumerate(todas_linhas):
        for idx, celula in enumerate(linha):
            if not celula: continue
            valor = str(celula).strip().upper()

            if valor in ["PARAMETRO", "COMANDO"]:
                col_param = idx
            elif valor in ["DESCRICAO", "DESCRIÇÃO", "AJUSTES GLOBAIS"]:
                col_desc = idx
            elif valor == "FAIXA DE AJUSTE":
                col_faixa = idx
            elif valor == "AJUSTE SUGERIDO":
                col_ajuste = idx

        if col_param is not None and col_ajuste is not None:
            linha_inicio_dados = i
            break
    
    # resolver esse print aqui
    if col_param is None or col_ajuste is None:
        print("Não foi possível localizar coluna 'PARAMETRO' ou 'AJUSTE' no arquivo")
        return []
    

    if linha_inicio_dados > 0:
        linha_anterior = todas_linhas[linha_inicio_dados - 1]
        for celula in linha_anterior:
            if celula and str(celula).strip():
                grupo_atual = str(celula).strip()
                break


    for i in range(linha_inicio_dados + 1, len(todas_linhas)):
        linha = todas_linhas[i]

        texto_linha_completa = " ".join([str(c).upper() for c in linha if c])
        
        # ÚLTIMO BLOCO DA PLANILHA
        if "ELABORAÇÃO E IMPLANTAÇÃO" in texto_linha_completa or "ELABORACAO E IMPLANTACAO" in texto_linha_completa:
            break
        
        raw_param = linha[col_param] if len(linha) > col_param else None
        valor_param = str(raw_param).strip() if raw_param else None

        if linha_metadados(valor_param, texto_linha_completa):
            continue

        raw_ajuste = linha[col_ajuste] if len(linha) > col_ajuste else None
        valor_ajuste = str(raw_ajuste).strip() if raw_ajuste is not None else None
        
        if valor_param and valor_param.upper() in ["PARAMETRO", "COMANDO"]:
            linha_anterior = todas_linhas[i - 1]
            for celula in linha_anterior:
                if celula and str(celula).strip():
                    grupo_atual = str(celula).strip()
                    break
            continue 

        if (not valor_param and not valor_ajuste) or (valor_param and not valor_ajuste):
            texto_encontrado = None
            for celula in linha:
                txt = str(celula).strip()
                if celula and txt and txt.upper() != "NONE":
                    texto_encontrado = txt
                    break
            
            if texto_encontrado:
                if texto_encontrado.upper() not in ["AJUSTES GLOBAIS", "DESCRIÇÃO"]:
                    grupo_atual = texto_encontrado
            continue

        if valor_param and valor_ajuste:
            if valor_ajuste.upper() in ["AJUSTE SUGERIDO", "VALOR", "AJUSTE"]:
                continue

            desc = ""
            if col_desc is not None and len(linha) > col_desc and linha[col_desc] is not None:
                desc = str(linha[col_desc]).strip()
            
            faixa = ""
            if col_faixa is not None and len(linha) > col_faixa and linha[col_faixa] is not None:
                faixa = str(linha[col_faixa]).strip()

            lista_parametros.append(
                ParametroReferencia(
                    grupo=grupo_atual,
                    parametro=valor_param,
                    descricao=desc,
                    faixa_ajuste=faixa,
                    ajuste_referencia=valor_ajuste,
                )
            )
    return lista_parametros

# LEITURA E EXTRAÇÃO DO ARQUIVO DA ORDEM DE AJUSTE
def leitura_oa(conteudo_arquivo: bytes, nome_arquivo: str) -> ArquivoOA:
    if not conteudo_arquivo:
        raise ValueError("Arquivo da OA vazio.")
    
    if len(conteudo_arquivo) > MAX_TAMANHO_XLSX_BYTES:
        raise ValueError(
            f"Arquivo OA excede o tamanho máximo permitido "
            f"({MAX_TAMANHO_XLSX_BYTES / 1024 / 1024:.0f} MB)."
        )
    
    arquivo = load_workbook(filename=BytesIO(conteudo_arquivo), data_only=True)
    ordem_ajuste = arquivo.worksheets[0]

    rele_tipo = formatar_nome(ordem_ajuste["D3"].value)
    subestacao = str(ordem_ajuste["D6"].value).strip()

    lista_parametros = leitura_parametros_oa(ordem_ajuste)
    arquivo.close()

    dados_oa = ArquivoOA(
        nome_arquivo=nome_arquivo,
        subestacao=subestacao,
        rele_tipo=rele_tipo,
        parametros=lista_parametros
    )

    return dados_oa



# EXTRAÇÃO DOS PARÂMETROS DA IED
def leitura_parametros_ied(texto: str) -> list[ParametroAtual]:
    lista_parametros = []

    for chave, valor in PADRAO_PARAMETRO_IED.findall(texto):
        if chave in CHAVES_IGNORADAS_IED:
            continue
        
        # REMOVER COMENTÁRIOS
        if "#" in valor:
            valor = valor.split("#")[0]

        lista_parametros.append(ParametroAtual(
            parametro=chave,
            valor_atual=valor.strip()
        ))

    return lista_parametros

# LEITURA E EXTRAÇÃO DO ARQUIVO DA IED
def leitura_ied(conteudo_arquivo: bytes, nome_arquivo: str) -> ArquivoIED:
    if not conteudo_arquivo:
        raise ValueError("Arquivo IED vazio")

    texto = conteudo_arquivo.decode("utf-8", errors="ignore")
    
    rele_tipo = "NÃO IDENTIFICADO"
    match_rele = re.search(r"FID=([^\s]+)", texto)
    if match_rele:
        rele_tipo = formatar_nome(match_rele.group(1))

    lista_parametros = leitura_parametros_ied(texto)

    dados_ied = ArquivoIED(
        nome_arquivo=nome_arquivo,
        rele_tipo=rele_tipo,
        parametros=lista_parametros
    )

    return dados_ied